import sqlite3
from datetime import datetime

import openpyxl
import psycopg2

from get_dop_inf import m_e_spisok, GP_spisok, host, user, password, db_name, port
from typing import NamedTuple


class Exits(NamedTuple):
    kust: str
    m_e: str
    exit_date: datetime
    GP: int
    RUO: bool
    SNPH: bool


class Enterances(NamedTuple):
    kust: str
    m_e: str
    Ist_stage: datetime
    IInd_stage: datetime
    GP: int
    RUO: bool
    SNPH: bool


def import_exits(start_string: int, finish_string: int) -> list:
    # """Импортируем данные из Экселя и вставляем в список"""
    wb = openpyxl.load_workbook('Движение_БУ.xlsx', data_only=True)
    ws = wb['Движение_БУ']
    Exits_: [list] = []
    for number_of_string in range(start_string, finish_string):
        kust = ws.cell(number_of_string, 4).value
        m_e = ws.cell(number_of_string, 5).value
        if m_e is not None and m_e not in m_e_spisok:
            print("В строке №", number_of_string, "ошибка в названии месторождения. Нужно исправить")
        exit_date = ws.cell(number_of_string, 6).value
        GP = ws.cell(number_of_string, 8).value
        if GP is not None and GP not in GP_spisok:
            print("В строке №", number_of_string, "ошибка в грузоподъемности. Нужно исправить")
        RUO = ws.cell(number_of_string, 9).value
        if RUO is not None and RUO != 0 and RUO != 1:
            print("В строке №", number_of_string, "ошибка в типе БР. Нужно исправить")
        SNPH = ws.cell(number_of_string, 10).value
        if SNPH is not None and SNPH != 0 and SNPH != 1:
            print("В строке №", number_of_string, "ошибка в SNPH. Нужно исправить")
        Exit_ = Exits(kust, m_e, GP, RUO, SNPH, exit_date)
        if Exit_.kust == Exit_.m_e == Exit_.GP is None:
            continue
        else:
            Exits_.append(Exit_)
    return Exits_


def export_exits(start_string: int, finish_string: int):
    # """Полученный список вставляем в БД"""
    Exits_ = import_exits(start_string, finish_string)
    connection = psycopg2.connect(
        host=host,
        user=user,
        password=password,
        database=db_name,
        port=port
    )
    connection.autocommit = True
    with connection.cursor() as cursor:
        for Exit_ in Exits_:
            cursor.execute("INSERT INTO exit_ VALUES (%s, %s, %s, %s, %s, %s);", Exit_)
    connection.close()


def import_enterances(start_string: int, finish_string: int) -> list:
    # """Импортируем данные из Экселя и вставляем в список"""
    wb = openpyxl.load_workbook('Движение_БУ.xlsx', data_only=True)
    ws = wb['Движение_БУ']
    Enterances_: [list] = []
    for number_of_string in range(start_string, finish_string):
        kust = ws.cell(number_of_string, 12).value
        m_e = ws.cell(number_of_string, 13).value
        if m_e is not None and m_e not in m_e_spisok:
            print("В строке №", number_of_string, "ошибка в названии месторождения. Нужно исправить")
        GP = ws.cell(number_of_string, 17).value
        if GP is not None and GP not in GP_spisok:
            print("В строке №", number_of_string, "ошибка в грузоподъемности. Нужно исправить")
        RUO = ws.cell(number_of_string, 19).value
        if RUO is not None and RUO != 0 and RUO != 1:
            print("В строке №", number_of_string, "ошибка в типе БР. Нужно исправить")
        SNPH = ws.cell(number_of_string, 20).value
        if SNPH is not None and SNPH != 0 and SNPH != 1:
            print("В строке №", number_of_string, "ошибка в SNPH. Нужно исправить")
        Ist_stage = ws.cell(number_of_string, 14).value
        IInd_stage = ws.cell(number_of_string, 15).value
        Enterance = Enterances(kust, m_e, GP, RUO, SNPH, Ist_stage, IInd_stage)
        if Enterance.kust == Enterance.m_e == Enterance.GP is None:
            continue
        else:
            Enterances_.append(Enterance)
    return Enterances_


def export_enterances(start_string: int, finish_string: int):
    # """Полученный список вставляем в БД"""
    Enterances_ = import_enterances(start_string, finish_string)
    connection = psycopg2.connect(
        host=host,
        user=user,
        password=password,
        database=db_name,
        port=port
    )
    connection.autocommit = True
    with connection.cursor() as cursor:
        for Enterance in Enterances_:
            cursor.execute("INSERT INTO enterance VALUES (%s, %s, %s, %s, %s, %s, %s);", Enterance)
    connection.close()
